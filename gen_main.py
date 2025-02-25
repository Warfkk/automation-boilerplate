from re import I
import sys
import os
from os import listdir
from os.path import isfile, join
import json
import openpyxl as xl
from settings import Settings
import shutil

# User Libs
from obj_lib.General import General
from obj_lib.alarm import Alarm
from obj_lib.asi import ASi
from unit_lib.unit_types import UnitTypes
from unit_lib.units_phases import UnitsPhases


class GenMain:
    """Main class called from UI,
    read data from excel and then execute sub-class functions
    """

    def __init__(self):
        self.s = Settings()
        self.user_settings = self.s.user_settings

        self.excel_path = self.user_settings["excel_path"]
        self.output_path = self.user_settings["output_path"]
        self.config_path = self.user_settings["config_path"]
        self.cm_output_path = os.path.join(self.output_path, "CMs")
        self.alarm_output_path = os.path.join(self.output_path, "Alarm")
        self.unit_output_path = os.path.join(self.output_path, "Units_Phases")
        self.plcinexcel = set()
        self.dict_list = []

        self.generate()

    def _open_gen_excel(self):
        try:
            wb = xl.load_workbook(self.excel_path, data_only=True)
        except FileNotFoundError as e:
            print(e)
            print("ERROR! Excel file not found, program will exit")
            sys.exit()
        else:
            self.wb = wb

    def copy_excel_data_to_dictionaries(self):
        """Open excel and read all relevant object-data to dict"""
        self._open_gen_excel()

        # Create all dictionaries, if enabled in settings
        if self.user_settings["DI_ENABLE"]:
            self.di_dict = self._obj_data_to_dict(
                self.s.DI_SHEETNAME, self.s.DI_START_INDEX, "di"
            )
            self.dict_list.append(self.di_dict)

        if self.user_settings["DO_ENABLE"]:
            self.do_dict = self._obj_data_to_dict(
                self.s.DO_SHEETNAME, self.s.DO_START_INDEX, "do"
            )
            self.dict_list.append(self.do_dict)

        if self.user_settings["VALVE_ENABLE"]:
            self.valve_dict = self._obj_data_to_dict(
                self.s.VALVE_SHEETNAME, self.s.VALVE_START_INDEX, "valve", config=True
            )
            self.dict_list.append(self.valve_dict)

        if self.user_settings["MOTOR_ENABLE"]:
            self.motor_dict = self._obj_data_to_dict(
                self.s.MOTOR_SHEETNAME, self.s.MOTOR_START_INDEX, "motor", config=True
            )
            self.dict_list.append(self.motor_dict)

        if self.user_settings["AI_ENABLE"]:
            self.ai_dict = self._obj_data_to_dict(
                self.s.AI_SHEETNAME, self.s.AI_START_INDEX, "ai", eng_var=True
            )
            self.dict_list.append(self.ai_dict)

        if self.user_settings["AO_ENABLE"]:
            self.ao_dict = self._obj_data_to_dict(
                self.s.AO_SHEETNAME, self.s.AO_START_INDEX, "ao", eng_var=True
            )
            self.dict_list.append(self.ao_dict)

        if self.user_settings["PID_ENABLE"]:
            self.pid_dict = self._obj_data_to_dict(
                self.s.PID_SHEETNAME, self.s.PID_START_INDEX, "pid", eng_var=True
            )
            self.dict_list.append(self.pid_dict)

        if self.user_settings["SUM_ENABLE"]:
            self.sum_dict = self._obj_data_to_dict(
                self.s.SUM_SHEETNAME,
                self.s.SUM_START_INDEX,
                "sum",
                eng_var=True,
                volumeperpulse=True,
            )
            self.dict_list.append(self.sum_dict)

        if self.user_settings["ALARM_ENABLE"]:
            self.alarm_dict = self._obj_data_to_dict(
                self.s.ALARM_SHEETNAME,
                self.s.ALARM_START_INDEX,
                "alarm",
                generic_alarm=True,
            )
            self.dict_list.append(self.alarm_dict)

        if self.user_settings["ASI_ENABLE"]:
            self.asi_dict = self._obj_data_to_dict(
                self.s.ASI_SHEETNAME, self.s.ASI_START_INDEX, "asi", asi=True
            )
            self.dict_list.append(self.asi_dict)

        if self.user_settings["UNITS_ENABLE"]:
            self.unit_phase_list = self._unit_data_to_list(self.s.UNIT_SHEETNAME)

    def _obj_data_to_dict(
        self,
        sheet,
        start_index,
        type,
        config=False,
        eng_var=False,
        volumeperpulse=False,
        generic_alarm=False,
        asi=False,
    ):
        """Read all object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f"ERROR! {sheet} sheet does not exist, program will exit"
            print(msg)
            sys.exit()

            # Loop header and set the corresponding variables to
            # the integer number
        for i in range(1, 20):
            cell = ws.cell(row=self.s.HEADER_ROW, column=i)
            cellval = str(cell.value)

            # If cell is empty (NoneType) - skip it
            if cellval is None:
                continue

            if self.s.COL_ID_NAME == cellval:
                column_id = i
                column_ioid = i
            if self.s.COL_IOID_NAME == cellval:
                column_ioid = i
            if self.s.COL_COMMENT_NAME == cellval:
                column_comment = i
            if self.s.COL_ALARM_GROUP_NAME == cellval:
                column_alarmgroup = i
            if self.s.COL_PLC_NAME == cellval:
                column_plc = i

            if type == "pid":
                if self.s.COL_IDPV_NAME == cellval:
                    column_idpv = i
                if self.s.COL_IDCV_NAME == cellval:
                    column_idcv = i

            if config:
                if self.s.COL_CONFIG_NAME == cellval:
                    column_config = i

            if volumeperpulse:
                if self.s.COL_VolumePerPulse_Name == cellval:
                    column_volumeperpulse = i

            if eng_var:
                if self.s.COL_ENG_UNIT_NAME == cellval:
                    column_eng_unit = i
                if self.s.COL_ENG_MIN_NAME == cellval:
                    column_eng_min = i
                if self.s.COL_ENG_MAX_NAME == cellval:
                    column_eng_max = i

            if generic_alarm:
                if self.s.COL_ALARM_PRIO_NAME == cellval:
                    column_alarm_prio = i
                if self.s.COL_ALARM_TEXT_NAME == cellval:
                    column_alarm_text = i

            if asi:
                if self.s.COL_ASI_ADDR_NAME == cellval:
                    column_asi_addr = i
                if self.s.COL_ASI_MASTER_NAME == cellval:
                    column_asi_master = i

        if self.s.debug_level > 0:
            print("SHEET:", sheet)
            print("\t", "column_id:", column_id)
            print("\t", "column_ioid:", column_ioid)
            print("\t", "column_comment:", column_comment)
            print("\t", "column_alarmgroup:", column_alarmgroup)
            print("\t", "column_plc:", column_plc)
            if config:
                print("\t", "column_config:", column_config)
            if volumeperpulse:
                print("\t", "column_volumeperpulse:", column_volumeperpulse)
            if eng_var:
                print("\t", "column_eng_unit:", column_eng_unit)
                print("\t", "column_eng_min:", column_eng_min)
                print("\t", "column_eng_max:", column_eng_max)
            if generic_alarm:
                print("\t", "column_alarm_prio:", column_alarm_prio)
                print("\t", "column_alarm_text:", column_alarm_text)
            if asi:
                print("\t", "column_asi_addr:", column_asi_addr)
                print("\t", "column_asi_master:", column_asi_master)
            if type == "pid":
                print("\t", "column_idpv:", column_idpv)
                print("\t", "column_idcv:", column_idcv)

        # Loop through object list and add key-value pairs to object dict
        # then append each object-dict to list
        obj_list = []
        index = start_index
        for i in range(self.s.ROW, ws.max_row + 1):
            # Break if we get a blank ID cell
            cell_id = ws.cell(row=i, column=column_id)
            cell_ioid = ws.cell(row=i, column=column_ioid)
            cell_comment = ws.cell(row=i, column=column_comment)
            cell_alarmgroup = ws.cell(row=i, column=column_alarmgroup)
            cell_plc = ws.cell(row=i, column=column_plc)

            if cell_id.value is None:
                break

            # Always insert these key-value pairs
            obj = {
                "type": type,
                "id": cell_id.value,
                "ioid": cell_ioid.value,
                "comment": cell_comment.value,
                "index": index,
                "alarmgroup": cell_alarmgroup.value,
                "plc": cell_plc.value,
            }

            # Add conditional key-value pairs
            if config:
                cell_config = ws.cell(row=i, column=column_config)
                obj["config"] = cell_config.value

            if volumeperpulse:
                cell_volumeperpulse = ws.cell(row=i, column=column_volumeperpulse)
                obj["volumeperpulse"] = cell_volumeperpulse.value

            if eng_var:
                cell_eng_unit = ws.cell(row=i, column=column_eng_unit)
                obj["eng_unit"] = cell_eng_unit.value
                cell_eng_min = ws.cell(row=i, column=column_eng_min)
                obj["eng_min"] = cell_eng_min.value
                cell_eng_max = ws.cell(row=i, column=column_eng_max)
                obj["eng_max"] = cell_eng_max.value

            if generic_alarm:
                cell_alarm_text = ws.cell(row=i, column=column_alarm_text)
                obj["alarm_text"] = cell_alarm_text.value
                cell_alarm_prio = ws.cell(row=i, column=column_alarm_prio)
                obj["alarm_prio"] = cell_alarm_prio.value

            if asi:
                cell_asi_addr = ws.cell(row=i, column=column_asi_addr)
                obj["asi_addr"] = cell_asi_addr.value
                cell_asi_master = ws.cell(row=i, column=column_asi_master)
                obj["asi_master"] = cell_asi_master.value

            if type == "pid":
                cell_idpv = ws.cell(row=i, column=column_idpv)
                obj["idpv"] = cell_idpv.value
                cell_idcv = ws.cell(row=i, column=column_idcv)
                obj["idcv"] = cell_idcv.value

            obj_list.append(obj)
            index += 1

        for obj in obj_list:
            self.plcinexcel.add(obj["plc"])

        return obj_list

    def _pars_data_to_dict(self, sheet, start_index, type, config=False):
        """Read all object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f"ERROR! {sheet} sheet does not exist, program will exit"
            print(msg)
            sys.exit()

            # Loop header and set the corresponding variables to
            # the integer number
        for i in range(1, 20):
            cell = ws.cell(row=self.s.HEADER_ROW, column=i)
            cellval = str(cell.value)

            # If cell is empty (NoneType) - skip it
            if cellval is None:
                continue

            if "Name" == cellval:
                column_name = i
            if "TYPE" == cellval:
                column_type = i
            if "OFFSET" == cellval:
                column_offset = i
            if "DESC" == cellval:
                column_desc = i
            if "DataType" == cellval:
                column_datatype = i
            if "DB" == cellval:
                column_db = i
            if "PLC" == cellval:
                column_plc = i

        if self.s.debug_level > 0:
            print("SHEET:", sheet)
            print("\t", "column_name:", column_name)
            print("\t", "column_type:", column_type)
            print("\t", "column_offset:", column_offset)
            print("\t", "column_desc:", column_desc)
            print("\t", "column_datatype:", column_datatype)
            print("\t", "column_db:", column_db)
            print("\t", "column_plc:", column_plc)

        # Loop through object list and add key-value pairs to object dict
        # then append each object-dict to list
        obj_list = []
        index = start_index
        for i in range(self.s.ROW, ws.max_row + 1):
            # Break if we get a blank ID cell
            cell_name = ws.cell(row=i, column=column_name)
            cell_type = ws.cell(row=i, column=column_type)
            cell_offset = ws.cell(row=i, column=column_offset)
            cell_desc = ws.cell(row=i, column=column_desc)
            cell_datatype = ws.cell(row=i, column=column_datatype)
            cell_db = ws.cell(row=i, column=column_db)
            cell_plc = ws.cell(row=i, column=column_plc)

            if cell_name.value is None:
                break

            # Always insert these key-value pairs
            obj = {
                "type": type,
                "name": cell_name.value,
                "type": cell_type.value,
                "offset": cell_offset.value,
                "desc": cell_desc.value,
                "datatype": cell_datatype.value,
                "db": cell_db.value,
                "plc": cell_plc.value,
            }

            obj_list.append(obj)
            index += 1

        return obj_list

    def _unit_data_to_list(self, sheet):
        """Read all unit object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f"ERROR! {sheet} sheet does not exist, program will exit"
            print(msg)
            sys.exit()

            # Loop header and set the corresponding variables to
            # the integer number
        for i in range(1, 10):
            cell = ws.cell(row=self.s.UNIT_HEADER_ROW, column=i)
            cellval = str(cell.value)

            # If cell is empty (NoneType) - skip it
            if cellval is None:
                continue

            if self.s.COL_ID_NAME == cellval:
                column_id = i
            elif self.s.COL_TYPE_NAME == cellval:
                column_type = i
            elif self.s.COL_PLC_NAME == cellval:
                column_plc = i
            elif self.s.COL_ALARM_GROUP_NAME == cellval:
                column_hmi_group = i

        if self.s.debug_level > 0:
            print("UNIT UNITSHEET:", sheet)
            print("\t", "UNIT column_id:", column_id)
            print("\t", "UNIT column_type:", column_type)
            print("\t", "UNIT column_plc:", column_plc)
            print("\t", "UNIT column_hmi_group:", column_hmi_group)

        unit_phase_list = []
        #  loop over the objects in sheet
        mem_unit = None
        mem_plc = None

        for i in range(self.s.UNIT_ROW, ws.max_row + 1):
            #  Create cell references
            cell_id = ws.cell(row=i, column=column_id)
            cell_type = ws.cell(row=i, column=column_type)
            cell_plc = ws.cell(row=i, column=column_plc)
            cell_hmi_group = ws.cell(row=i, column=column_hmi_group)

            # Break if we get a blank ID cell
            if cell_id.value is None:
                break

            def _is_valid_unit_type(in_type):
                for type in UnitTypes:
                    if in_type == type.value:
                        return True
                return False

            def _is_unit(in_type):
                if "Unit" in in_type:
                    return True
                else:
                    return False

            is_valid_unit_type = _is_valid_unit_type(cell_type.value)
            is_unit = _is_unit(cell_type.value)
            is_phase = not is_unit

            if is_unit:
                #  Remembers var from unit, in that way less
                #  duplicate data in excel
                mem_unit = cell_id.value
                mem_plc = cell_plc.value
                mem_hmi_group = cell_hmi_group.value
                parent = None
            else:
                parent = mem_unit

            # Create object dict, always with these key-value pairs
            obj = {
                "is_valid_unit_type": is_valid_unit_type,
                "is_unit": is_unit,
                "is_phase": is_phase,
                "parent": parent,
                "type": cell_type.value,
                "id": cell_id.value,
                "plc": mem_plc,
                "hmi_group": mem_hmi_group,
            }

            unit_phase_list.append(obj)

        return unit_phase_list

    def create_subdirs(self):
        """Create all subdirectiories beyond output path"""
        dirs = [self.s.TIA_DIR, self.s.INTOUCH_DIR, self.s.SQL_DIR]
        for dir in dirs:
            newdir = os.path.join(self.output_path, dir)
            if not os.path.exists(newdir):
                os.makedirs(newdir)

    def create_subdirsplc(self):
        """Create all subdirectiories beyond output path"""
        for plc in self.plcinexcel:
            newdir = os.path.join(self.output_path, self.s.TIA_DIR)
            newdir = os.path.join(newdir, plc)
            if not os.path.exists(newdir):
                os.makedirs(newdir)

    def get_config_from_config_path(self):
        """Load config from .json file"""
        json_file = os.path.join(self.config_path, "config_type.json")
        with open(json_file, "r") as (f):
            json_var = json.load(f)
            self.config_type = json_var["type"]
            print(f"Config Type={self.config_type}")

    @staticmethod
    def _print_disabled_in_settings(prefix):
        print(f"{prefix} not generated, disabled in settings file")

    def _combine_it_files(self):
        """
        combines all it files in folder and chops off first lines
        in files other than the first
        """

        for folder in listdir(self.output_path):
            filename = "ALL_" + folder + "_IT.csv"
            path = os.path.join(self.output_path, folder, self.s.INTOUCH_DIR)
            outfile = os.path.join(path, filename)

            if os.path.exists(outfile):
                os.remove(outfile)

            if os.path.exists(path):
                file_list = [f for f in listdir(path) if isfile(join(path, f))]

                with open(outfile, "w", encoding="cp1252") as wf:
                    for file_index, file in enumerate(file_list):
                        with open(
                            os.path.join(path, file), "r", encoding="cp1252"
                        ) as rf:
                            for line_index, line in enumerate(rf):
                                # Skip first line header if it's not the first file
                                if file_index > 0 and line_index <= 0:
                                    continue
                                wf.write(line)

    def _combine_sql_files(self):
        """
        combines all it files in folder and chops off first lines
        in files other than the first
        """

        for folder in listdir(self.output_path):
            filename = "ALL_" + folder + "_SQL.csv"
            path = os.path.join(self.output_path, folder, self.s.SQL_DIR)
            outfile = os.path.join(path, filename)

            if os.path.exists(outfile):
                os.remove(outfile)

            if os.path.exists(path):
                file_list = [f for f in listdir(path) if isfile(join(path, f))]

                with open(outfile, "w", encoding="cp1252") as wf:
                    for file_index, file in enumerate(file_list):
                        with open(
                            os.path.join(path, file), "r", encoding="cp1252"
                        ) as rf:
                            for line_index, line in enumerate(rf):
                                # Skip first line header if it's not the first file
                                if file_index > 0 and line_index <= 0:
                                    continue
                                wf.write(line)

    @staticmethod
    def _parse_s7_db_addr(in_db_addr):
        #  Format expected e.g. DB9001.DBX12.0, DB9001.DBB12, DB9001.DBW12, DB9001.DBD12
        splitted = in_db_addr.split(".")

        if len(splitted) == 3:
            # three elements == two dots == DB9001.DBX0.0 format
            db, dbx, bit = splitted
        elif len(splitted) == 2:
            #  two elements == one dot == any other format specified
            db, dbx = splitted

        remove_chars = ["D", "B", "X", "W", "D"]

        for char in remove_chars:
            dbx = dbx.replace(char, "")

        db_offset = int(dbx)

        return db, db_offset

    def generate(self):
        print("Version", self.s.version)

        # Remove output folder if it exists
        if os.path.exists(self.output_path):
            shutil.rmtree(self.output_path)

        self.copy_excel_data_to_dictionaries()

        if self.s.debug_level > 0:
            for dict in self.dict_list:
                for obj in dict:
                    print(obj)

        if self.user_settings["UNITS_ENABLE"] and self.s.debug_level > 0:
            for _ in self.unit_phase_list:
                print(_)

        self.get_config_from_config_path()

        if not self.user_settings["VALVE_ENABLE"]:
            self._print_disabled_in_settings("Valve")
        else:
            General(self, self.output_path, self.valve_dict, self.config_path, "Valves")

        if not self.user_settings["MOTOR_ENABLE"]:
            self._print_disabled_in_settings("Motor")
        else:
            General(self, self.output_path, self.motor_dict, self.config_path, "Motors")

        if not self.user_settings["DI_ENABLE"]:
            self._print_disabled_in_settings("DI")
        else:
            General(
                self, self.output_path, self.di_dict, self.config_path, "DigitalInputs"
            )

        if not self.user_settings["DO_ENABLE"]:
            self._print_disabled_in_settings("DO")
        else:
            General(
                self, self.output_path, self.do_dict, self.config_path, "DigitalOutputs"
            )
        if not self.user_settings["AI_ENABLE"]:
            self._print_disabled_in_settings("AI")
        else:
            General(
                self, self.output_path, self.ai_dict, self.config_path, "AnalogInputs"
            )

        if not self.user_settings["AO_ENABLE"]:
            self._print_disabled_in_settings("AO")
        else:
            General(
                self, self.output_path, self.ao_dict, self.config_path, "AnalogOutputs"
            )

        if not self.user_settings["PID_ENABLE"]:
            self._print_disabled_in_settings("PID")
        else:
            General(self, self.output_path, self.pids_dict, self.config_path, "PIDs")

        if not self.user_settings["SUM_ENABLE"]:
            self._print_disabled_in_settings("SUM")
        else:
            General(self, self.output_path, self.sum_dict, self.config_path, "SUMs")

        if not self.user_settings["ALARM_ENABLE"]:
            self._print_disabled_in_settings("Alarm")
        else:
            Alarm(
                self,
                self.output_path,
                self.alarm_dict,
                self.config_path,
                config_type=self.config_type,
            )

        if not self.user_settings["ASI_ENABLE"]:
            self._print_disabled_in_settings("ASi")
        else:
            ASi(
                self,
                self.output_path,
                self.asi_dict,
                self.config_path,
                config_type=self.config_type,
            )

        if not self.user_settings["UNITS_ENABLE"]:
            self._print_disabled_in_settings("Units_Phases")
        else:
            UnitsPhases(
                self,
                self.output_path,
                self.unit_phase_list,
                self.config_path,
                config_type=self.config_type,
            )

        self._combine_it_files()
        self._combine_sql_files()
